import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook

def descombinar_celdas(ws):
    for merged_cell in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_cell))
        
def guardar_resultados(ruta_archivo_excel, resultado, diferencia, valores, objetivo):
    
    df_resultados = pd.DataFrame({
        'Cantidad de Valores': [len(valores)],
        'Valor Objetivo': [objetivo],
        'Sumatoria de Resultados': [sum(resultado)],
        'Diferencia': [diferencia]
    })
    
    # Cargar el archivo Excel
    libro = load_workbook(ruta_archivo_excel)
    hoja_actual = libro.active
    
    # Eliminar combinaciones de celdas
    descombinar_celdas(hoja_actual)
    
    # Leer la hoja actual en un DataFrame
    df = pd.read_excel(ruta_archivo_excel)
    
    # Obtener la columna 'A' como una lista
    columna_a = df.iloc[:, 0].tolist()
    
    # Crear la columna 'Resultados'
    resultados_columna = []
    resultado_set = set(resultado)  # Convertir a set para búsqueda rápida
    
    for valor in columna_a:
        if valor in resultado:
            resultados_columna.append(valor)  # Extraer y usar el primer valor de resultado
            resultado.remove(valor)
        else:
            resultados_columna.append(0)  # Asignar 0 si el valor no está en resultado
    
    # Crear el DataFrame 'df_proceso' con las modificaciones
    df_proceso = pd.DataFrame({
        'Valores': columna_a,
        'Resultados': resultados_columna
    })
    
    # Colores
    morado_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
    blanco_font = Font(color="FFFFFF", bold=True)
    
    # Escribir los DataFrames en el archivo Excel
    with pd.ExcelWriter(ruta_archivo_excel, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=hoja_actual.title, index=False)
        df_proceso.to_excel(writer, sheet_name='Proceso', index=False)
        df_resultados.to_excel(writer, sheet_name='Resultados', index=False)
        
        hoja_proceso = writer.sheets['Proceso']
        hoja_resultados = writer.sheets['Resultados']
        
        for cell in hoja_proceso[1]:
            cell.fill = morado_fill
            cell.font = blanco_font
        
        for cell in hoja_resultados[1]:
            cell.fill = morado_fill
            cell.font = blanco_font
        
        for col in hoja_proceso.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            hoja_proceso.column_dimensions[column].width = adjusted_width
            
        for col in hoja_resultados.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            hoja_resultados.column_dimensions[column].width = adjusted_width
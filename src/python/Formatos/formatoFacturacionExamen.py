import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from Directories.Directory import DirectoryFormatoFacturacionExamen

def get_postulado_row(postulado):
    try:
        return {
            "Proveedor de servicio": str(postulado['proveedor']),
            "Cliente": str(postulado['cliente']),
            "Cargo": str(postulado['cargo']),
            "Nombres y apellidos": str(postulado['nombre']),
            "Numero de documento": str(postulado['documento']),
            "Requisicion": str(postulado['requisicion']),
            "Tipo de examen": str(postulado['tipo_examen']),
            "Fecha del examen": str(postulado['fecha_examen']),
            "Centro medico": str(postulado['centro_medico']),
            "Sede medica": str(postulado['sede_medica']),
            "Examenes": str(postulado['examenes'])
        }
    except ValueError as e:
        print(f"Error procesando datos del postulado: {e}")

# Objeto de borde
border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

# Estilos para la primera fila
def style_header(header):
    font = Font(size=12, bold=True, color='FFFFFF')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    fill = PatternFill(fill_type="solid", fgColor="C81C1C")
    
    for cell in header:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        cell.fill = fill
        
# Configurar las filas y columnas del excel
def auto_size_columns_and_rows(ws):
    
    # Ajustar ancho de las columnas
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Ajustar el alto y el borde de las filas según el contenido
    for row in ws.iter_rows(min_row=6, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
        
        for cell in row:
            try:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
            except:
                pass

def create_workbook_and_sheet(rows):
    wb = Workbook()
    ws = wb.active

    column_titles = list(rows[0].keys())
    ws.append(column_titles)
    style_header(ws[1])

    for row in rows:
        ws.append(list(row.values()))

    auto_size_columns_and_rows(ws)

    return wb, ws
    
# Procesar json
def process_json_data(data):
    try:
        json_object = json.loads(data)
        
        postulados = json_object['data']['postulados']
        
        rows = [get_postulado_row(postulado) for postulado in postulados]
        
        wb, ws = create_workbook_and_sheet(rows)
        
        # File
        path = DirectoryFormatoFacturacionExamen
        path = "./formato.xlsx"
        wb.save(path)
        return path

    except json.JSONDecodeError as e:
        print(f"Error al decodificar JSON: {e}")
    except Exception as e:
        print(f"Error inesperado: {e}")
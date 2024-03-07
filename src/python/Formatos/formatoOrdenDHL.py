import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from Directories.Directory import DirectoryFormatoFacturacionExamen

def get_postulado_row(orden):
    try:
        return {
            "TIPO DE VINCULACION": str(orden['tipo_vinculacion']),
            "NIT": str(orden['nit']),
            "EMPRESA TEMPORAL": str(orden['temporal']),
            "CEDULA": str(orden['cc']),
            "APELLIDOS Y NOMBRES": str(orden['nombres']),
            "CODIGO EMPRESA": str(orden['codigo_empresa']),
            "EMPRESA ASOCIADA": str(orden['empresa']),
            "CARGO": str(orden['cargo']),
            "CATEGORIA": str(orden['categoria']),
            "EMAIL EMPLEADO": str(orden['email']),
            "FECHA INGRESO": str(orden['fecha_ingreso']),
            "FECHA RETIRO": str(orden['fecha_retiro']),
            "SALARIO": str(orden['salario']),
            "JOB FUNCTION": str(orden['job_function']),
            "COD TYPE OF WORK": str(orden['cod_type_work']),
            "TYPE OF WORK": str(orden['type_work']),
            "AT": str(orden['at']),
            "AT NAME": str(orden['at_name']),
            "SUB": str(orden['sub']),
            "CU": str(orden['cu']),
            "CU NAME": str(orden['cu_name']),
            "COST CENTER": str(orden['cost_center']),
            "COST CENTER NAME": str(orden['cost_center_name']),
            "DIRECCION OFICINA": str(orden['direccion_oficina']),
            "POBLACION": str(orden['poblacion']),
            "CODE SECTOR": str(orden['code_sector']),
            "SECTOR": str(orden['sector']),
            "GENERO": str(orden['genero']),
            "CEDULA JEFE": str(orden['cedula_jefe']),
            "NOMBRES JEFE": str(orden['nombres_jefe']),
            "%ARP": str(orden['arp']),
            "FECHA DE NACIMIENTO": str(orden['fecha_nacimiento']),
            "EPS": str(orden['eps']),
            "AFP": str(orden['afp']),
            "FONDO DE CESANTIAS": str(orden['fondo_cesantias']),
            "CAJA DE COMPENSACION": str(orden['caja_compensacion']),
            "# DE CUENTA BANCARIA": str(orden['cta_bancaria']),
            "TIPO DE CUENTA": str(orden['tipo_cuenta']),
            "BANCO": str(orden['banco']),
            "TELEFONO": str(orden['telefono']),
            "CELULAR": str(orden['celular']),
            "ESTADO DE CONTRATO": str(orden['estado_contrato']),
            "TIPO DE CONTRATO": str(orden['tipo_contrato']),
            "MOTIVO DEL INGRESO": str(orden['motivo_ingreso']),
            "MOTIVO DEL RETIRO": str(orden['motivo_retiro']),
            "PANTALON": str(orden['pantalon']),
            "CAMISA / CAMISETA /OVEROL": str(orden['camisa']),
            "BOTAS / TENIS": str(orden['botas']),
            "CHAQUETA": str(orden['chaqueta'])            
        }
    except ValueError as e:
        print(f"Error procesando datos de la orden: {e}")

# Objeto de borde
border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

# Estilos para la primera fila
def style_header(header):
    font = Font(size=12, bold=True, color='FFFFFF')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    fill = PatternFill(fill_type="solid", fgColor="B45F06")
    
    for cell in header:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        cell.fill = fill
        
# Configurar las filas y columnas del excel
def auto_size_columns_and_rows(ws):
    
    # Ajustar ancho de las columnas
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Ajustar el alto y el borde de las filas según el contenido
    for row in ws.iter_rows(min_row=6, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
        
        for cell in row:
            try:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
            except:
                pass

def create_workbook_and_sheet(rows):
    wb = Workbook()
    ws = wb.active

    column_titles = list(rows[0].keys())
    ws.append(column_titles)
    style_header(ws[1])

    for row in rows:
        ws.append(list(row.values()))

    auto_size_columns_and_rows(ws)

    return wb, ws
    
# Procesar json
def process_json_data(data):
    try:
        json_object = json.loads(data)
        
        ordenes = json_object['data']['orden']
        
        rows = [get_postulado_row(orden) for orden in ordenes]
        
        wb, ws = create_workbook_and_sheet(rows)
        
        # File
        path = DirectoryFormatoFacturacionExamen
        path = "./formato.xlsx"
        wb.save(path)
        return path

    except json.JSONDecodeError as e:
        print(f"Error al decodificar JSON: {e}")
    except Exception as e:
        print(f"Error inesperado: {e}")
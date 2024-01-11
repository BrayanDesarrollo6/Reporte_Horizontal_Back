import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from Directories.Directory import DirectoryFormatoOrdenIngreso

# Objeto de borde
border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

# Procesar json
def process_json_data(data):
    try:
        json_object = json.loads(data)
        
        requisicion = json_object['data']['requisicion']
        orden_ingreso = json_object['data']['orden_ingreso']
        beneficios_prestacionales = json_object['data']['beneficios_prestacionales']
        beneficios_no_prestacionales = json_object['data']['beneficios_no_prestacionales']
        postulados = json_object['data']['postulados']
        documentos_especiales = json_object['data']['documentos_especiales']

        rows = [get_postulado_row(postulado, requisicion, orden_ingreso, beneficios_prestacionales, beneficios_no_prestacionales, documentos_especiales) for postulado in postulados]
        
        wb, ws = create_workbook_and_sheet(rows)
        
        add_footer(ws)
        add_signature_block(ws)
        add_notes(ws)
        add_header(ws)
        add_header_2(ws)
        add_images(ws)
        
        # File
        path = DirectoryFormatoOrdenIngreso
        wb.save(path)
        return path

    except json.JSONDecodeError as e:
        print(f"Error al decodificar JSON: {e}")
    except Exception as e:
        print(f"Error inesperado: {e}")

def get_postulado_row(postulado, requisicion, orden_ingreso, beneficios_prestacionales, beneficios_no_prestacionales, documentos_especiales):
    try:
        return {
            "Tipo de documento": str(postulado['tipo_doc']),
            "No. de identificación": str(postulado['numero_doc']),
            "Nombres y apellidos": str(postulado['nombre']),
            "Empresa temporal": str(requisicion['proveedor_servicios']),
            "Empresa usuaria": str(requisicion['cliente']),
            "Cargo": str(requisicion['cargo']),
            "Fecha de ingreso": str(orden_ingreso['fecha_ingreso']),
            "Ciudad donde desarrollará labores": str(requisicion['ciudad']),
            "Sitio de trabajo": str(orden_ingreso['sitio_trabajo']),
            "Centro de costo": str(orden_ingreso['centro_costo']),
            "Tipo de riesgo": str(orden_ingreso['nivel_riesgo']),
            "Horario en que se prestará el servicio": str(requisicion['horario']),
            "Lugar, hora y persona con la que debe presentarse": str(orden_ingreso['sitio_presentacion']),
            "Salario asignado": str(requisicion['salario']),
            "Otros conceptos (Prestacional) $": str(", ".join(str(b['grupo']) for b in beneficios_prestacionales)),
            "Tipo de concepto (Prestacional)": str(", ".join(str(b['concepto']) for b in beneficios_prestacionales)),
            "Valor (Prestacional)": str(", ".join(str(b['valor']) for b in beneficios_prestacionales)),
            "Otros conceptos no salariales (No prestacional) $": str(", ".join(str(b['grupo']) for b in beneficios_no_prestacionales)),
            "Tipo de concepto (No prestacional)": str(", ".join(str(b['concepto']) for b in beneficios_no_prestacionales)),
            "Valor (No prestacional)": str(", ".join(str(b['valor']) for b in beneficios_no_prestacionales)),
            "Documentos especiales (Otro si)": str(documentos_especiales)
        }
    except ValueError as e:
        print(f"Error procesando datos del postulado: {e}")

def create_workbook_and_sheet(rows):
    wb = Workbook()
    ws = wb.active

    column_titles = list(rows[0].keys())
    ws.append(column_titles)
    style_header(ws[1])

    for row in rows:
        ws.append(list(row.values()))

    ws.insert_cols(1)
    for i in range(3):
        ws.insert_rows(1)
    auto_size_columns_and_rows(ws)

    return wb, ws

# Estilos para la primera fila
def style_header(header):
    font = Font(size=12, bold=True, color='FFFFFF')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    fill = PatternFill(fill_type="solid", fgColor="C81C1C")
    
    for cell in header:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        cell.fill = fill
        
# Configurar las filas y columnas del excel
def auto_size_columns_and_rows(ws):
    
    # Ajustar ancho de las columnas
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
        
        # Establecer ancho fijo para las columnas desde I hasta V
        for col_num in range(9, 24):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 15
            ws.column_dimensions['A'].width = 3
            ws.column_dimensions['M'].width = 30
            ws.column_dimensions['Q'].width = 30
            ws.column_dimensions['T'].width = 30
    
    # Ajustar el alto y el borde de las filas según el contenido
    for row in ws.iter_rows(min_row=5, min_col=2, max_col=ws.max_column, max_row=ws.max_row):
        
        for cell in row:
            try:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
            except:
                pass

def add_footer(ws):
    max_row = ws.max_row

    for _ in range(2):
        ws.insert_rows(max_row + 1)

    ws.merge_cells(f"B{max_row + 2}:{get_column_letter(ws.max_column)}{max_row + 2}")
    footer_cell = ws.cell(row=max_row + 2, column=2)
    style_footer(footer_cell)
    value = "Autorizacion Empresa usuaria"
    footer_cell.value = value
    ws.row_dimensions[max_row + 2].height = 50

    for i in range(21):
        footer_cell = ws.cell(row=max_row + 2, column=2 + i)
        footer_cell.border = border

# Estilos para las ultimas filas         
def style_footer(cell):
    font = Font(size=12, bold=True, color='FFFFFF')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    fill = PatternFill(fill_type="solid", fgColor="C81C1C")
    cell.font = font
    cell.alignment = alignment
    cell.fill = fill
    cell.border = border

def add_empty_rows(ws, start_row, num_rows):
    for i in range(num_rows):
        ws.insert_rows(start_row + 1)

def add_signature_block(ws):
    max_row = ws.max_row
    add_empty_rows(ws, max_row, 2)

    signature_cell = ws.cell(row=max_row + 2, column=2)
    value = "Nombres y apellidos"
    signature_cell.value = value
    style_signature(signature_cell)

    ws.merge_cells(f"C{max_row + 2}:E{max_row + 2}")
    for i in range(4):
        signature_cell = ws.cell(row=max_row + 2, column=2 + i)
        signature_cell.border = border

    signature_cell = ws.cell(row=max_row + 2, column=6)
    value = "Cargo"
    signature_cell.value = value
    style_signature(signature_cell)

    ws.merge_cells(f"G{max_row + 2}:J{max_row + 2}")
    for i in range(5):
        signature_cell = ws.cell(row=max_row + 2, column=6 + i)
        signature_cell.border = border

    ws.row_dimensions[max_row + 2].height = 50

# Estilos para las firmas
def style_signature(cell):
    font = Font(size=12, bold=True, color='FFFFFF')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    fill = PatternFill(fill_type="solid", fgColor="C81C1C")
    cell.font = font
    cell.alignment = alignment
    cell.fill = fill
    cell.border = border
    
def add_notes(ws):
    max_row = ws.max_row
    add_empty_rows(ws, max_row, 2)

    ws.merge_cells(f"B{max_row + 2}:{get_column_letter(ws.max_column)}{max_row + 2}")
    note_cell = ws.cell(row=max_row + 2, column=2)
    style_notes(note_cell)
    value = "Nota: Se tomará como firma y autorización válida si la orden es emitida por un correo electrónico de dominio de la empresa usuaria \n\n" \
            "La versión impresa de este documento es una copia no controlada. \n" \
            "Este material es propiedad intelectual de TEMPOENLACE - HQ5 queda prohibida su reproducción total o parcial"
    note_cell.value = value
    ws.row_dimensions[max_row + 2].height = 100

    for i in range(21):
        note_cell = ws.cell(row=max_row + 2, column=2 + i)
        note_cell.border = border
        
# Estilos para las notas 
def style_notes(cell):
    font = Font(size=12, bold=True, color='000000')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    
def add_header(ws):
    ws.merge_cells(f"D2:S2")
    title_cell = ws.cell(row=2, column=4)
    style_headers(title_cell)
    value = "ORDEN DE INGRESO PROCESO MASIVO"
    title_cell.value = value
    ws.row_dimensions[2].height = 80

    for i in range(16):
        title_cell = ws.cell(row=2, column=4 + i)
        title_cell.border = border
        
# Estilos para el Header
def style_headers(cell):
    font = Font(size=24, bold=True, color='000000')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.font = font
    cell.alignment = alignment
    cell.border = border

def add_header_2(ws):
    ws.merge_cells(f"T2:V2")
    title_cell = ws.cell(row=2, column=20)
    style_headers_2(title_cell)
    fecha_actual = datetime.now()
    formato_fecha = fecha_actual.strftime("%d-%m-%Y")
    value = f"Código: MM-AT-F-02\nFecha: {formato_fecha}\nVersión: 2\nPágina: 1/1"
    title_cell.value = value

    for i in range(3):
        title_cell = ws.cell(row=2, column=20 + i)
        title_cell.border = border

# Estilos para el Header 2
def style_headers_2(cell):
    font = Font(size=12, bold=True, color='000000')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.font = font
    cell.alignment = alignment
    cell.border = border

def add_images(ws):
    image_urls = [
        "https://raw.githubusercontent.com/HQ5SAS/Recursos_Publicos/main/recursos/hq5_logo.png",
        "https://raw.githubusercontent.com/HQ5SAS/Recursos_Publicos/main/recursos/tempoenlace1_logo.png"
    ]

    for idx, image_url in enumerate(image_urls, start=2):
        img = get_resized_image(image_url, ws.column_dimensions[get_column_letter(idx)].width, ws.row_dimensions[2].height)
        ws.add_image(img, f"{get_column_letter(idx)}2")
        image_cell = ws.cell(row=2, column=idx)
        image_cell.border = border
        
def get_resized_image(image_url, target_width, target_height):
    response = requests.get(image_url)
    img_stream = BytesIO(response.content)
    img = Image(img_stream)

    aspect_ratio = img.width / img.height
    new_width = target_width
    new_height = new_width / aspect_ratio

    if new_height > target_height:
        new_height = target_height
        new_width = new_height * aspect_ratio

    img.width = new_width * 6.5
    img.height = new_height * 6.5
    
    return img
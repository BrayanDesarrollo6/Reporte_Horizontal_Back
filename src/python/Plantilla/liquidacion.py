import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def liquidacion(d_temporal,d_cliente,d_empleados,d_periodo,d_concepto,d_valor,d_unidades,d_id_temporal,d_id_cliente,d_id_concepto,d_id_periodo):
    
    d_temporal = d_temporal.replace(" ", "%20")
    d_cliente = d_cliente.replace(" ", "%20")
    
    if(d_empleados != "AMBOS"):
        d_empleados = d_empleados.replace(" ", "%20")
        URL = "https://creatorapp.zohopublic.com/hq5colombia/compensacionhq5/xls/Reporte_de_Empleados_General/Tnq3UKDyevz83bx7VCdZBV1PYJ5HNVuHRKMSSvpTAR18R6aCOk32HZPNsWnRRHButMC9jAamCfBW5g6Js5z1GwNz0a34D18jvUwY?Estado_Trabajador=" + d_empleados + "&Empresa=" + d_cliente + "&Temporal.TEMPORAL=" + d_temporal
        df = pd.read_excel(URL)
    else:
        URL = "https://creatorapp.zohopublic.com/hq5colombia/compensacionhq5/xls/Reporte_de_Empleados_General/Tnq3UKDyevz83bx7VCdZBV1PYJ5HNVuHRKMSSvpTAR18R6aCOk32HZPNsWnRRHButMC9jAamCfBW5g6Js5z1GwNz0a34D18jvUwY?Empresa=" + d_cliente + "&Temporal.TEMPORAL=" + d_temporal
        df = pd.read_excel(URL)

    if df.empty:
        print("No existe registro")
        return

    rows = []
    for i in range(len(df)):
        FilaAgregar = {
            "Empresa": str(d_id_cliente),
            "Nombre Concepto": str(d_id_concepto),
            "Documento": str(df.iloc[i]['ID System']),
            "Contrato": str(df.iloc[i]['ID System']),
            "Unidades hora": int(d_unidades),
            "Valor": float(d_valor),
            "Informacion": str(df.iloc[i]['Numero de Documento']) + " - " + str(df.iloc[i]['Nombre Completo']) + " - " + str(df.iloc[i]['N° de Contrato']) + " - " + str(df.iloc[i]['Estado Trabajador'])
        }
        rows.append(FilaAgregar)

    wb = Workbook()
    ws = wb.active

    column_titles = list(FilaAgregar.keys())
    ws.append(column_titles)

    font = Font(size=12, bold=False, color='000000')
    alignment = Alignment(horizontal='center')
    fill = PatternFill(fill_type="solid", fgColor="8ECD82")
    
    for cell in ws[1]:
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill

    for row in rows:
        ws.append(list(row.values()))

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    relleno_columna1 = PatternFill(start_color='8ECD82', end_color='8ECD82', fill_type='solid')
    rango_columna1 = ws['G:G']
    
    for celda in rango_columna1:
        celda.fill = relleno_columna1   

    path = "./src/database/Liquidacion.xlsx"
    wb.save(path)
    return path
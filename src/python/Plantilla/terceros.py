import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from Directories.Directory import DirectoryPlantillaTerceros

def terceros(d_temporal,d_cliente,d_empleados,d_concepto,d_fecha,d_tipo,d_valor,d_realizacion_descuento,d_n_cuotas,d_modo_pago,d_estado_des_total):

    d_temporal = d_temporal.replace(" ", "%20")
    d_cliente = d_cliente.replace(" ", "%20")
    
    if(d_empleados != "AMBOS"):
        d_empleados = d_empleados.replace(" ", "%20")
        URL = "https://creatorapp.zohopublic.com/hq5colombia/compensacionhq5/xls/Reporte_de_Empleados_General/Tnq3UKDyevz83bx7VCdZBV1PYJ5HNVuHRKMSSvpTAR18R6aCOk32HZPNsWnRRHButMC9jAamCfBW5g6Js5z1GwNz0a34D18jvUwY?Estado_Trabajador=" + d_empleados + "&Empresa=" + d_cliente + "&Temporal.TEMPORAL=" + d_temporal
        df = pd.read_excel(URL)
    else:
        URL = "https://creatorapp.zohopublic.com/hq5colombia/compensacionhq5/xls/Reporte_de_Empleados_General/Tnq3UKDyevz83bx7VCdZBV1PYJ5HNVuHRKMSSvpTAR18R6aCOk32HZPNsWnRRHButMC9jAamCfBW5g6Js5z1GwNz0a34D18jvUwY?Empresa=" + d_cliente + "&Temporal.TEMPORAL=" + d_temporal
        df = pd.read_excel(URL)

    if df.empty:
        print("No existe registro")
        return

    rows = []
    for i in range(len(df)):
        FilaAgregar = {
            "Documento": int(df.iloc[i]['Numero de Documento']),
            "Contrato": int(df.iloc[i]['N° de Contrato']),
            "Nombre Concepto": str(d_concepto),
            "Tipo": str(d_tipo),
            "Modo de pago": str(d_modo_pago),
            "Valor Total": float(d_valor),
            "Fecha de inicio": str(d_fecha),
            "Cuando se realiza el descuento": str(d_realizacion_descuento),
            "Número de cuotas": int(d_n_cuotas),
            "Estado descuento total": str(d_estado_des_total)
        }
        rows.append(FilaAgregar)

    wb = Workbook()
    ws = wb.active

    column_titles = list(FilaAgregar.keys())
    ws.append(column_titles)

    font = Font(size=12, bold=False, color='000000')
    alignment = Alignment(horizontal='center')
    fill = PatternFill(fill_type="solid", fgColor="8ECD82")
    
    for cell in ws[1]:
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill

    for row in rows:
        ws.append(list(row.values()))

    opciones = ['Prestamo', 'Abono voluntario']
    opciones_2 = ['Cada periodo', 'Cada primer periodo', 'Cada segundo periodo']
    validacion_col1 = DataValidation(type="list", formula1=f'"{",".join(opciones)}"')
    validacion_col2 = DataValidation(type="list", formula1=f'"{",".join(opciones_2)}"')

    rango_celdas_col1 = ws['D2:D' + str(ws.max_row)]
    rango_celdas_col2 = ws['H2:H' + str(ws.max_row)]

    for fila in rango_celdas_col1:
        for celda in fila:
            ws.add_data_validation(validacion_col1)
            validacion_col1.add(celda)

    for fila in rango_celdas_col2:
        for celda in fila:
            ws.add_data_validation(validacion_col2)
            validacion_col2.add(celda)

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 16

    relleno_columna1 = PatternFill(start_color='8ECD82', end_color='8ECD82', fill_type='solid')
    relleno_columna2 = PatternFill(start_color='8ECD82', end_color='8ECD82', fill_type='solid')
    rango_columna1 = ws['E:E']
    rango_columna2 = ws['J:J']

    for celda in rango_columna1:
        celda.fill = relleno_columna1

    for celda in rango_columna2:
        celda.fill = relleno_columna2

    path = DirectoryPlantillaTerceros
    wb.save(path)
    return path
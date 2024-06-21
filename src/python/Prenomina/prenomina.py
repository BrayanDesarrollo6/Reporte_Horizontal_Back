import requests
import json
import time
from datetime import datetime
from os import remove
import os
import pandas as pd
from xlsxwriter import Workbook
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from io import BytesIO
import sys
import cv2
import asyncio

#Importar funciones para obtener BD
from accessToken import funcionesGenerales

# Objeto de borde
border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

# reemplazar acentos
def normalize(s):
    replacements = (
        ("á", "a"),("é", "e"),("í", "i"),("ó", "o"),("ú", "u"),
        ("Á", "A"),("É", "E"),("Í", "I"),("Ó", "O"),("Ú", "U")
    )
    for a, b in replacements:
        s = s.replace(a, b).replace(a.upper(), b.upper())
    return s

# reemplazar caracteres
def replacement(name_company):
    name_company = name_company.replace(".", "")
    name_company = name_company.replace("-", "_")
    name_company = name_company.replace("–", "_")
    name_company = name_company.replace("—", "_")
    return name_company

def add_header(ws):
    ws.merge_cells(f"D2:AU2")
    title_cell = ws.cell(row=2, column=4)
    style_headers(title_cell)
    value = "SOLICITUD DE PRENOMINA"
    title_cell.value = value
    ws.row_dimensions[2].height = 80

    for i in range(47):
        title_cell = ws.cell(row=2, column=4 + i)
        title_cell.border = border
        
# Estilos para el Header
def style_headers(cell):
    font = Font(size=24, bold=True, color='000000')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.font = font
    cell.alignment = alignment
    cell.border = border

def add_header_2(ws):
    # Asigna el valor a la celda antes de la fusión
    title_cell = ws.cell(row=2, column=48)  # AU es la columna 40
    value = "Código: F-GOP-008\nFecha: 22-07-2019\nVersión: 2\nPágina: 1/1"
    title_cell.value = value
    style_headers_2(title_cell)

    # Fusión de celdas después de asignar el valor
    ws.merge_cells("AV2:AX2")

    # Aplica estilos a las celdas fusionadas
    for i in range(3):
        title_cell = ws.cell(row=2, column=47 + i)
        title_cell.border = border

# Estilos para el Header 2
def style_headers_2(cell):
    font = Font(size=12, bold=True, color='000000')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.font = font
    cell.alignment = alignment
    cell.border = border

def add_images(ws):
    image_urls = [
        "https://raw.githubusercontent.com/HQ5SAS/Recursos_Publicos/main/recursos/hq5_logo.png",
        "https://raw.githubusercontent.com/HQ5SAS/Recursos_Publicos/main/recursos/tempoenlace1_logo.png"
    ]

    for idx, image_url in enumerate(image_urls, start=2):
        img = get_resized_image(image_url, ws.column_dimensions[get_column_letter(idx)].width, ws.row_dimensions[2].height)
        ws.add_image(img, f"{get_column_letter(idx)}2")
        image_cell = ws.cell(row=2, column=idx)
        image_cell.border = border
    
def add_info_customer(ws,df_info):
    # ws.merge_cells(f"D2:AT2")
    title_cell = ws.cell(row=4, column=2)
    style_info_customer(title_cell)
    title_cell.value = "CLIENTE"
    #PERIODICIDAD DE PAGO
    title_cell = ws.cell(row=5, column=2)
    style_info_customer(title_cell)
    title_cell.value = "PERIODO DE PAGO"
    
    ws.merge_cells(f"C4:E4")
    title_cell = ws.cell(row=4, column=3)
    style_info_customer(title_cell)
    title_cell.value = str(df_info.iloc[0]['Empresa'])
    periodicidad="QUINCENAL"
    if(str(df_info.iloc[0]['Tipo de Perido']) == "3"):
        periodicidad="MENSUAL"
    ws.merge_cells(f"C5:E5")
    title_cell = ws.cell(row=5, column=3)
    style_info_customer(title_cell)
    title_cell.value = periodicidad
    
    
    ws.row_dimensions[2].height = 80

    for i in range(2):
        title_cell = ws.cell(row=4, column=4 + i)
        title_cell.border = border
        title_cell = ws.cell(row=5, column=4 + i)
        title_cell.border = border

def style_info_customer(cell):
    font = Font(size=12, bold=True, color='000000')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.font = font
    cell.alignment = alignment
    cell.border = border
def get_resized_image(image_url, target_width, target_height):
    response = requests.get(image_url)
    img_stream = BytesIO(response.content)
    img = Image(img_stream)

    aspect_ratio = img.width / img.height
    new_width = target_width
    new_height = new_width / aspect_ratio

    if new_height > target_height:
        new_height = target_height
        new_width = new_height * aspect_ratio

    img.width = new_width * 6.5
    img.height = new_height * 6.5
    
    return img
# Configurar las filas y columnas del excel
def auto_size_columns_and_rows(ws):
    
    # Ajustar ancho de las columnas
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
        
        # Establecer ancho fijo para las columnas desde I hasta V
        for col_num in range(9, 24):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 15
            ws.column_dimensions['A'].width = 3
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['M'].width = 30
            ws.column_dimensions['Q'].width = 30
            ws.column_dimensions['T'].width = 30
    
    # Ajustar el alto y el borde de las filas según el contenido
    for row in ws.iter_rows(min_row=8, min_col=2, max_col=ws.max_column, max_row=ws.max_row):
        
        for cell in row:
            try:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            except:
                pass
# Estilos para la primera fila
def style_header(header):
    font = Font(size=10, bold=True, color='000000')
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    fill = PatternFill(fill_type="solid", fgColor="FF5FE6")
    
    for cell in header:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        cell.fill = fill

def add_suplmentarios(ws):
    contador_ = 0
    for x in range(10,19):
        title_cell = ws.cell(row=1, column=x)
        # style_headers(title_cell)
        value = suplementarios[contador_]["name"]
        title_cell.value = value
        ws.row_dimensions[2].height = 80
        itle_cell = ws.cell(row=1, column=x)
        title_cell.border = border
        # subname
        title_cell = ws.cell(row=2, column=x)
        # style_headers(title_cell)
        value = suplementarios[contador_]["subname"]
        title_cell.value = value
        ws.row_dimensions[2].height = 80
        itle_cell = ws.cell(row=2, column=x)
        title_cell.border = border
        # Porcentaje
        title_cell = ws.cell(row=3, column=x)
        # style_headers(title_cell)
        value = suplementarios[contador_]["porc"]
        title_cell.value = value
        ws.row_dimensions[2].height = 80
        itle_cell = ws.cell(row=3, column=x)
        title_cell.border = border
        contador_ += 1
    return ws

def add_devengos(ws):
    contador_ = 0
    # ws.merge_cells(start_row=1, start_column=19, end_row=1, end_column=20)
    for x in range(19,25):
        title_cell = ws.cell(row=1, column=x)
        # style_headers(title_cell)
        value = "DEVENGOS"
        title_cell.value = value
        title_cell = ws.cell(row=2, column=x)
        # style_headers(title_cell)
        value = devengos[contador_]["name"]
        title_cell.value = value
        itle_cell = ws.cell(row=2, column=x)
        title_cell.border = border
        # subname
        title_cell = ws.cell(row=3, column=x)
        # style_headers(title_cell)
        value = devengos[contador_]["subname"]
        title_cell.value = value
        itle_cell = ws.cell(row=3, column=x)
        title_cell.border = border
        contador_ += 1
    return ws
  
def add_deducciones(ws):
    contador_ = 0
    for x in range(25,31):
        title_cell = ws.cell(row=1, column=x)

        value = "DEDUCCIONES"
        title_cell.value = value
        title_cell = ws.cell(row=2, column=x)
        # style_headers(title_cell)
        value = deducciones[contador_]["name"]
        title_cell.value = value
        itle_cell = ws.cell(row=2, column=x)
        title_cell.border = border
        # subname
        title_cell = ws.cell(row=3, column=x)
        # style_headers(title_cell)
        value = deducciones[contador_]["subname"]
        title_cell.value = value
        itle_cell = ws.cell(row=3, column=x)
        title_cell.border = border
        contador_ += 1
    return ws
  
def add_ausentismos(ws):
    contador_ = 0
    title_cell = ws.cell(row=1, column=31)
    value = "AUSENTISMO"
    for x in range(31,48,3):
        
        title_cell.value = value
        title_cell = ws.cell(row=2, column=x)
        value = ausentismos[contador_]["name"]
        title_cell.value = value
        title_cell.border = border
        # Fecha inicial
        title_cell = ws.cell(row=3, column=x)
        value = "Fecha inicial"
        title_cell.value = value
        title_cell.border = border
        # Fecha final
        title_cell = ws.cell(row=3, column=x+1)
        value = "Fecha final"
        title_cell.value = value
        title_cell.border = border
        # Días
        title_cell = ws.cell(row=3, column=x+2)
        value = "Días"
        title_cell.value = value
        title_cell.border = border
        contador_ += 1
    # OBSERVACIONES
    title_cell = ws.cell(row=1, column=49)
    value = "OBSERVACIONES"
    title_cell.value = value
    title_cell.border = border
    return ws
  
def create_workbook_and_sheet(df):
    wb = Workbook()
    ws = wb.active

    # Insertar los encabezados después del merge
    column_titles = list(df.columns)
    ws.append(column_titles)
    ws = add_suplmentarios(ws)
    ws = add_devengos(ws)
    ws = add_deducciones(ws)
    ws = add_ausentismos(ws)
    style_header(ws[1])
    style_header(ws[2])
    style_header(ws[3])

    for _ in range(7):
        ws.insert_rows(1)
    # Insertar los datos del DataFrame
    for _, row in df.iterrows():
        ws.append(list(row))

    # Insertar columna adicional
    ws.insert_cols(1)
    # Merge para informacion empleado
    for x in range(2, 11):
        ws.merge_cells(start_row=8, start_column=x, end_row=10, end_column=x)
    # MERGE PAR AOBSERVACIONES
    for x in range(2, 11):
        ws.merge_cells(start_row=8, start_column=50, end_row=10, end_column=50)
    # Merge para devengos
    ws.merge_cells(f"T8:Y8")
    # Merge para deducciones
    ws.merge_cells(f"Z8:AE8")
    # Merge para Ausentismos
    ws.merge_cells(f"AF8:AW8")
    # Para cada ausentismos
    # INCAPACIDAD EPS
    ws.merge_cells(f"AF9:AH9")
    # INCAPACIDAD ARL
    ws.merge_cells(f"AI9:AK9")
    # CALAMIDAD
    ws.merge_cells(f"AL9:AN9")
    # PERMISO NO REMUNERADO
    ws.merge_cells(f"AO9:AQ9")
    # SUSPENSION DICIPLINARIA
    ws.merge_cells(f"AR9:AT9")
    # AUSENCIA NO JUSTIFICADA
    ws.merge_cells(f"AU9:AW9")
    # Ajustar automáticamente el tamaño de las columnas según el contenido
    auto_size_columns_and_rows(ws)

    return wb, ws

def generar_row(df_horizontal,df_prenomina):
    FilaAgregar = {}
    ##Informacion general inicial
    FilaAgregar["ID EMPLEADO"] = df_horizontal.iloc[0]['ID empleado']
    FilaAgregar["CONTRATO"] = df_horizontal.iloc[0]['Numero de Contrato']
    FilaAgregar["NIT"] = df_horizontal.iloc[0]['Numero de Documento']
    FilaAgregar["NOMBRES"] = df_horizontal.iloc[0]['Nombre Completo']
    FilaAgregar["CARGO"] = df_horizontal.iloc[0]['Cargo Contratado']
    FilaAgregar['DEPENDENCIA\nO\nCENTRO DE COSTO'] = df_horizontal.iloc[0]['Centro de Costo']
    fecha_inicial = df_horizontal.iloc[0]['Fecha Ingreso']
    if str(fecha_inicial) != "NaT":
        # fecha = datetime.strptime(fecha_inicial, '%Y-%m-%d %H:%M:%S')
        fecha = fecha_inicial.date()
        FilaAgregar["FECHA INGRESO"] = fecha
    else:
        FilaAgregar["FECHA INGRESO"] = ""
    fecha_final = df_horizontal.iloc[0]['Fecha Retiro SS']
    if str(fecha_final) != "NaT" and str(fecha_final) != "":
        fechafin = fecha_final.date()
        FilaAgregar["FECHA FINAL"] = fechafin
    else:
        FilaAgregar["FECHA FINAL"] = ""
    FilaAgregar["BÁSICO"] = '${:,.2f}'.format(float(df_horizontal.iloc[0]['Salario Básico']))
    FilaAgregar["DÍAS"] = ""

    
    df_prenomina = pd.concat([df_prenomina,pd.DataFrame.from_records([FilaAgregar])],ignore_index=True)
    return df_prenomina  


def procesar(prenomina_zoho):
    # Dataframe final para obtener los indices de las primeras columnas 
    global Empresa_
    Horizontal = pd.DataFrame()
    Horizontal_heads_end = pd.DataFrame()
    #ELIMINAR REPETIDOS
    prenomina_zoho.drop_duplicates(subset =['Numero de Contrato'], keep="last", inplace=True)
    Horizontal_heads_end = prenomina_zoho
    # Obtener empleados
    Contrato  = prenomina_zoho['Numero de Contrato'].unique().tolist()
    for contratoX in Contrato:
        Valores = prenomina_zoho['Numero de Contrato'] == contratoX
        ContratoPos = prenomina_zoho[Valores]
        Horizontal = generar_row(ContratoPos,Horizontal)
        
    #Nombre documento
    Empresa_ = prenomina_zoho.iloc[0]['Empresa']
    NombreDocumento = "Prenomina_" + Empresa_
    # Normalizar nombre del documento
    NombreDocumento = normalize(NombreDocumento)
    NombreDocumento = replacement(NombreDocumento)
    heads = Horizontal.columns.values
    FilaAgregar = {}
    Validador = False

    wb, ws = create_workbook_and_sheet(Horizontal)
    add_header(ws)
    add_header_2(ws)
    add_images(ws)
    add_info_customer(ws,prenomina_zoho)

    # Puede generar error
    wb.save("./src/"+NombreDocumento+".xlsx")
    funcionesGenerales().updatedata("./src/"+NombreDocumento+".xlsx",IDregistro_)
  
def procesar_prenomina(data):
    # -----------------------------
    global IDregistro_,Periodo,suplementarios,devengos,deducciones,ausentismos
    json_object = json.loads(data)
    Periodo = json_object['data']['periodo']
    IDregistro_ = json_object['data']['id_registro']
    
    #Nombre para columnas
    suplementarios = [
      {
        "name": "Horas Extra Diurna",
        "subname": "HED",
        "porc": "125%"
      },
      {
        "name": "Hora Extra Nocturna",
        "subname": "HEN",
        "porc": "175%"
      },
      {
        "name": "Hora Extra Dominical Diurna",
        "subname": "HEFD",
        "porc": "200%"
      },
      {
        "name": "Hora Extra Dominical Nocturna",
        "subname": "HEFN",
        "porc": "250%"
      },
      {
        "name": "Dominical Sin Compensatorio",
        "subname": "DSC/F",
        "porc": "175%"
      },
      {
        "name": "Dominical Con Compensatorio",
        "subname": "DCC",
        "porc": "75%"
      },
      {
        "name": "Recargo Nocturno",
        "subname": "RN",
        "porc": "35%"
      },
      {
        "name": "Recargo Nocturno Dominical",
        "subname": "RND",
        "porc": "210%"
      },
      {
        "name": "Recargo Nocturno Dominical Compensado",
        "subname": "RNDC",
        "porc": "110%"
      }
    ]
    #Nombre para columnas devengos
    devengos = [
        {
          "name":"CONCEPTO SALARIAL",
          "subname":""
        },
        {
          "name":"VALOR SALARIAL",
          "subname":""
        },
        {
          "name":"CONCEPTO NO SALARIAL",
          "subname":""
        }
        ,
        {
          "name":"VALOR NO SALARIAL",
          "subname":""
        }
        ,
        {
          "name":"OTRO CONCEPTO NO SALARIAL",
          "subname":""
        }
        ,
        {
          "name":"VALOR OTRO CONCEPTO NO SALARIAL",
          "subname":""
        }
      ]
    #Nombre para columnas devengos
    deducciones = [
        {
          "name":"CONCEPTO",
          "subname":"DESCUENTOS"
        },
        {
          "name":"VALOR",
          "subname":""
        },
        {
          "name":"CONCEPTO",
          "subname":"DESCUENTOS"
        },
        {
          "name":"VALOR",
          "subname":""
        },
        {
          "name":"CONCEPTO",
          "subname":"DESCUENTOS"
        },
        {
          "name":"VALOR",
          "subname":""
        }
      ]
    #Nombre para columnas devengos
    ausentismos = [
    {
      "name": "INCAPACIDAD EPS"
    },
    {
      "name": "INCAPACIDAD ARL"
    },
    {
      "name": "CALAMIDAD"
    },
    {
      "name": "PERMISO NO REMUNERADO"
    },
    {
      "name": "SUSPENSION DISCIPLINARIA"
    },
    {
      "name": "AUSENCIA NO JUSTIFICADA"
    }
  ]
    # RECORRER LAS PRESTACIONES SOCIALES
    URL = "https://creatorapp.zohopublic.com/hq5colombia/compensacionhq5/xls/Prenomina/WWjRAOJ2MGyyNGd5BxdvwApYGzgq5A9AQ5Q6bUmpsTQvWTMJE4qE5MyKnY4KKPXneurq8RnTZ2O698AO8N2KQ7Fa7qt4hpwSet0K?Temporal_op=30&Periodo=" + Periodo
    df = pd.read_excel(URL)
    df1 = pd.DataFrame(df)
    if(df1.empty):
        print("No existe registro")
    else:
        print("Existe y se esta procesando")
        procesar(df1) 
        
def main():
    info_received = sys.argv[1]
    procesar_prenomina(info_received) 
if __name__ == "__main__":
    main()
    # asyncio.run(main())
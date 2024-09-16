import json
import pandas as pd
import sys
from Formatos.fileUpload import Updatedata as upload_zoho_file
from Access.Getaccess import obtener_access_token
import openpyxl

def get_df(data):
    date_from = data["fecha_desde"]
    date_to = data["fecha_hasta"]
    
    op_param ="lp_examenes_programados.FECHA_A_REALIZAR_EL_EXAMEN_op=58"
    date_param = "lp_examenes_programados.FECHA_A_REALIZAR_EL_EXAMEN="+date_from+";"+date_to
    
    params = [date_param,op_param]
    
    if('clientes' in data):
        clientes = data["clientes"]
        clientes = ','.join(clientes).replace(' ','+')
        clientes_param =  "lp_examenes_programados.EMPRESA_USUARIA="+"["+clientes+"]"
        params.append(clientes_param)
    
    url_params = '&'.join(params)
    
    url = "https://creatorapp.zohopublic.com/hq5colombia/hq5/xls/Profesiograma_Programacion_individual_Report/FjtXFPONhKfDuhOagEY0qMKWfRnMRXXu9Z3W1mWdmQZwNF1Qq1qHRhfA82YVT9DFPZ2S9RWHp8EsuKgEC9Q8r1WShRZ5zqYwb7C7?"
    url = url + url_params
    
    df = pd.read_excel(url, parse_dates=['FECHA'])
    return df

def tranform_df(df):
    default_columns = df.shape[1]-2
    df['FECHA'] = df['FECHA'].dt.strftime('%d-%m-%Y')
    df['EXAMEN MEDICO'] = df['EXAMEN MEDICO'].apply(lambda x: x.replace('[', '').replace(']', '').replace('"', '') if isinstance(x, str) else '')
    nuevo_df = df.pivot_table(index='ID PROGRAMACION', columns='Examen medico', values='Valor', fill_value=0, aggfunc='first')
    nuevo_df = pd.merge(df.drop(columns=['Examen medico', 'Valor']).drop_duplicates(),nuevo_df, on='ID PROGRAMACION')

    columnas_examenes = nuevo_df.columns[default_columns:]
    nuevo_df['Total Exámenes'] = nuevo_df[columnas_examenes].sum(axis=1)
    
    total = [''] * len(nuevo_df.columns)
    total[len(total)-1] = nuevo_df['Total Exámenes'].sum()
    nuevo_df.loc[len(nuevo_df.index)] = total
    return nuevo_df

def df_to_excel(df):
    file_name = "reporte_facturacion.xlsx"  
    df.to_excel(file_name, index=False)
    return file_name

def apply_excel_styles(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    header_color = '024715'
    
    #header color
    for fila in sheet.iter_rows(min_row=1, max_row=1):
        for cell in fila:
            cell.fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    
    #cell width
    for column in sheet.columns:
        longitud_maxima = 0
        for cell in column:
            if cell.value:
                longitud_maxima = max(longitud_maxima, len(str(cell.value)))
        ajuste = (longitud_maxima + 2) * 1.2  # Ajuste arbitrario
        sheet.column_dimensions[column[0].column_letter].width = ajuste

    wb.save(file_name)

def main():
    req_data = json.loads(sys.argv[1])
    record_id = req_data["record_id"]
    report = req_data["report"]
    field_name = req_data["field_name"]    
    
    df = get_df(req_data["data"])
    df = tranform_df(df)
    file_name = df_to_excel(df)
    apply_excel_styles(file_name)
    
    access_token = obtener_access_token()
    upload_zoho_file(access_token,file_name,record_id,report,field_name)

if __name__ == "__main__":
    main()